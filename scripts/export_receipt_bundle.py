from __future__ import annotations

import html
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ACCENT = "245F52"
LIGHT = "F7F3EC"


def money(value):
    if value is None:
        return ""
    return f"{float(value):.2f}".replace(".", ",")


def text(value, fallback=""):
    if value is None:
        return fallback
    return str(value)


def load_payload(path: str):
    return json.loads(Path(path).read_text("utf-8"))


def resolve_participant(payload):
    participant_id = payload.get("participantId")
    distribution = payload.get("distributionState") or {}
    for participant in distribution.get("participants", []):
        if participant.get("id") == participant_id:
            return participant
    return None


def summarize_distribution(distribution):
    participants = distribution.get("participants", [])
    entries = distribution.get("entries", [])
    participant_map = {participant["id"]: {**participant, "assignments": [], "total": 0} for participant in participants}
    remaining_entries = []

    for entry in entries:
        assignments = entry.get("assignments", [])
        assigned_amount = 0
        assigned_quantity = 0

        for assignment in assignments:
          participant = participant_map.get(assignment.get("participantId"))
          if not participant:
              continue
          participant["assignments"].append({
              "entryName": entry.get("name"),
              **assignment,
          })
          participant["total"] += float(assignment.get("amount") or 0)
          assigned_amount += float(assignment.get("amount") or 0)
          assigned_quantity += float(assignment.get("quantity") or 0)

        remaining_total = round(max(0.0, float(entry.get("lineTotal") or 0) - assigned_amount), 2)
        remaining_quantity = round(max(0.0, float(entry.get("quantity") or 0) - assigned_quantity), 3)
        if remaining_total > 0.001 or remaining_quantity > 0.001:
            remaining_entries.append({
                "name": entry.get("name"),
                "remainingTotal": remaining_total,
                "remainingQuantity": remaining_quantity,
                "unitPrice": float(entry.get("unitPrice") or 0),
            })

    for participant in participant_map.values():
        participant["total"] = round(participant["total"], 2)

    remaining_total = round(sum(item["remainingTotal"] for item in remaining_entries), 2)
    return list(participant_map.values()), remaining_entries, remaining_total


def style_header(cells):
    fill = PatternFill("solid", fgColor=ACCENT)
    font = Font(color="FFFFFF", bold=True)
    for cell in cells:
        cell.fill = fill
        cell.font = font


def autosize(sheet):
    for column_cells in sheet.columns:
        values = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(values + [10]) + 2, 40)


def safe_sheet_name(name, existing_names, fallback="Medlem"):
    cleaned = text(name, fallback).strip() or fallback
    cleaned = "".join("-" if char in '[]:*?/\\' else char for char in cleaned)
    cleaned = cleaned[:31].rstrip() or fallback
    candidate = cleaned
    index = 2

    while candidate in existing_names:
        suffix = f" {index}"
        candidate = f"{cleaned[: max(1, 31 - len(suffix))]}{suffix}".rstrip()
        index += 1

    existing_names.add(candidate)
    return candidate


def pdf_cell(value, style):
    content = html.escape(text(value)).replace("\n", "<br/>")
    return Paragraph(content or "&nbsp;", style)


def wrap_table_rows(rows, style):
    wrapped = []
    for index, row in enumerate(rows):
        if index == 0:
            wrapped.append(row)
            continue
        wrapped.append([pdf_cell(cell, style) for cell in row])
    return wrapped


def build_registered_xlsx(payload, output_path):
    result = payload["result"]
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Registrert kostnad"
    summary.append(["Felt", "Verdi"])
    style_header(summary[1])
    summary.append(["Butikk / Restaurant", result.get("merchantName") or ""])
    summary.append(["Dato", result.get("receiptDate") or ""])
    summary.append(["Tid", result.get("receiptTime") or ""])
    summary.append(["Valuta", result.get("currency") or ""])
    summary.append(["Kvittering total", result.get("grandTotal")])
    summary.append(["Sum registrerte varer", result.get("totals", {}).get("itemsTotal")])
    summary.append(["Avvik", result.get("totals", {}).get("difference")])

    lines = workbook.create_sheet("Linjer")
    lines.append(["Vare", "Antall", "Pris per vare", "Linjesum", "Rålinje"])
    style_header(lines[1])
    for item in result.get("lineItems") or result.get("items") or []:
        lines.append([
            item.get("name"),
            item.get("quantity"),
            item.get("unitPrice"),
            item.get("lineTotal"),
            item.get("rawLine"),
        ])

    for sheet in workbook.worksheets:
        autosize(sheet)

    workbook.save(output_path)


def build_distribution_xlsx(payload, output_path):
    distribution = payload["distributionState"]
    participants, remaining_entries, remaining_total = summarize_distribution(distribution)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Fordeling"
    overview.append(["Felt", "Verdi"])
    style_header(overview[1])
    overview.append(["Kvittering total", payload["result"].get("grandTotal")])
    overview.append(["Gjenstår å fordele", remaining_total])

    participant_sheet = workbook.create_sheet("Personer")
    participant_sheet.append(["Person", "Total"])
    style_header(participant_sheet[1])
    for participant in participants:
      participant_sheet.append([participant.get("name"), participant.get("total")])

    assignments = workbook.create_sheet("Fordelte linjer")
    assignments.append(["Person", "Vare", "Type", "Antall", "Andel %", "Belop"])
    style_header(assignments[1])
    for participant in participants:
        for assignment in participant.get("assignments", []):
            assignments.append([
                participant.get("name"),
                assignment.get("entryName"),
                assignment.get("type"),
                assignment.get("quantity"),
                assignment.get("percent"),
                assignment.get("amount"),
            ])

    remaining = workbook.create_sheet("Gjenstar")
    remaining.append(["Vare", "Gjenstaende antall", "Pris per vare", "Gjenstaende sum"])
    style_header(remaining[1])
    for entry in remaining_entries:
        remaining.append([
            entry.get("name"),
            entry.get("remainingQuantity"),
            entry.get("unitPrice"),
            entry.get("remainingTotal"),
        ])

    for sheet in workbook.worksheets:
        autosize(sheet)

    workbook.save(output_path)


def build_participant_xlsx(payload, output_path):
    distribution = payload["distributionState"]
    participant = resolve_participant(payload)
    participants, _, _ = summarize_distribution(distribution)
    participant_summary = next((item for item in participants if item["id"] == participant["id"]), None)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Personfordeling"
    sheet.append(["Person", participant.get("name")])
    sheet.append(["Total", participant_summary.get("total") if participant_summary else 0])
    sheet.append([])
    sheet.append(["Vare", "Type", "Antall", "Andel %", "Belop"])
    style_header(sheet[4])

    for assignment in participant_summary.get("assignments", []) if participant_summary else []:
        sheet.append([
            assignment.get("entryName"),
            assignment.get("type"),
            assignment.get("quantity"),
            assignment.get("percent"),
            assignment.get("amount"),
        ])

    autosize(sheet)
    workbook.save(output_path)


def build_event_xlsx(payload, output_path):
    summary_data = payload.get("summary") or {}
    receipts = payload.get("receipts") or []
    members = summary_data.get("members") or []
    workbook = Workbook()
    existing_sheet_names = {"Arrangementsoversikt", "Kvitteringer", "Linjer"}
    member_sheet_names = {}

    for member in members:
        member_sheet_names[member.get("id")] = safe_sheet_name(member.get("name"), existing_sheet_names)

    overview = workbook.active
    overview.title = "Arrangementsoversikt"
    overview.append(["Felt", "Verdi"])
    style_header(overview[1])
    overview.append(["Arrangement", payload.get("eventName") or "Arrangement"])
    overview.append(["Kvitteringer", summary_data.get("receiptCount") or 0])
    overview.append(["Totalt brukt", summary_data.get("totalSpent") or 0])
    overview.append(["Registrert betalt", summary_data.get("totalPaid") or 0])
    overview.append(["Fordelt pa personer", summary_data.get("totalUsed") or 0])
    overview.append(["Ufordelt rest", summary_data.get("unassignedTotal") or 0])
    overview.append(["Mangler betaler", summary_data.get("missingPayerCount") or 0])
    overview.append([])

    member_header_row = overview.max_row + 1
    overview.append(["Person", "Betalt", "Brukt", "Balanse", "Betalte kvitteringer", "Detaljark"])
    style_header(overview[member_header_row])
    if members:
        for member in members:
            overview.append([
                member.get("name"),
                member.get("paidTotal"),
                member.get("usedTotal"),
                member.get("balance"),
                member.get("paidReceiptCount"),
                member_sheet_names.get(member.get("id"), ""),
            ])
    else:
        overview.append(["Ingen medlemmer ennå", "", "", "", "", ""])

    receipt_sheet = workbook.create_sheet("Kvitteringer")
    receipt_sheet.append([
        "Kvittering",
        "Sted",
        "Dato",
        "Tid",
        "Betalt av",
        "Valuta",
        "Total",
        "Registrerte varer",
        "Avvik",
        "Fordelt",
        "Ufordelt",
    ])
    style_header(receipt_sheet[1])
    if receipts:
        for receipt in receipts:
            receipt_sheet.append([
                receipt.get("title"),
                receipt.get("merchantName"),
                receipt.get("receiptDate"),
                receipt.get("receiptTime"),
                receipt.get("paidByMemberName") or "Ikke valgt",
                receipt.get("currency"),
                receipt.get("grandTotal"),
                receipt.get("itemsTotal"),
                receipt.get("difference"),
                receipt.get("distributedTotal"),
                receipt.get("unassignedTotal"),
            ])
    else:
        receipt_sheet.append(["Ingen kvitteringer registrert", "", "", "", "", "", "", "", "", "", ""])

    line_sheet = workbook.create_sheet("Linjer")
    line_sheet.append([
        "Kvittering",
        "Sted",
        "Dato",
        "Betalt av",
        "Vare",
        "Antall",
        "Pris per vare",
        "Linjesum",
        "Ralinje",
    ])
    style_header(line_sheet[1])
    if receipts:
        for receipt in receipts:
            line_items = receipt.get("lineItems") or []
            if not line_items:
                line_sheet.append([
                    receipt.get("title"),
                    receipt.get("merchantName"),
                    receipt.get("receiptDate"),
                    receipt.get("paidByMemberName") or "Ikke valgt",
                    "Ingen linjer registrert",
                    "",
                    "",
                    "",
                    "",
                ])
                continue

            for item in line_items:
                line_sheet.append([
                    receipt.get("title"),
                    receipt.get("merchantName"),
                    receipt.get("receiptDate"),
                    receipt.get("paidByMemberName") or "Ikke valgt",
                    item.get("name"),
                    item.get("quantity"),
                    item.get("unitPrice"),
                    item.get("lineTotal"),
                    item.get("rawLine"),
                ])
    else:
        line_sheet.append(["Ingen kvitteringer registrert", "", "", "", "", "", "", "", ""])

    for member in members:
        member_sheet = workbook.create_sheet(member_sheet_names.get(member.get("id")) or "Medlem")
        member_sheet.append(["Felt", "Verdi"])
        style_header(member_sheet[1])
        member_sheet.append(["Person", member.get("name") or ""])
        member_sheet.append(["Betalt", member.get("paidTotal") or 0])
        member_sheet.append(["Brukt", member.get("usedTotal") or 0])
        member_sheet.append(["Balanse", member.get("balance") or 0])
        member_sheet.append(["Betalte kvitteringer", member.get("paidReceiptCount") or 0])
        member_sheet.append([])

        receipts_header_row = member_sheet.max_row + 1
        member_sheet.append([
            "Rolle",
            "Kvittering",
            "Sted",
            "Dato",
            "Tid",
            "Betalt av",
            "Kvittering total",
            "Brukt av medlem",
            "Betalt av medlem",
        ])
        style_header(member_sheet[receipts_header_row])

        receipt_summaries = member.get("receiptSummaries") or []
        if receipt_summaries:
            for receipt_summary in receipt_summaries:
                member_sheet.append([
                    receipt_summary.get("role"),
                    receipt_summary.get("receiptTitle"),
                    receipt_summary.get("merchantName"),
                    receipt_summary.get("receiptDate"),
                    receipt_summary.get("receiptTime"),
                    receipt_summary.get("paidByMemberName") or "Ikke valgt",
                    receipt_summary.get("receiptTotal"),
                    receipt_summary.get("memberUsedTotal"),
                    receipt_summary.get("memberPaidTotal"),
                ])
        else:
            member_sheet.append(["Ingen kvitteringsdetaljer", "", "", "", "", "", "", "", ""])

        member_sheet.append([])

        lines_header_row = member_sheet.max_row + 1
        member_sheet.append([
            "Kvittering",
            "Sted",
            "Dato",
            "Tid",
            "Betalt av",
            "Vare",
            "Type",
            "Antall",
            "Andel %",
            "Belop",
            "Pris per vare",
            "Linjesum",
        ])
        style_header(member_sheet[lines_header_row])

        line_assignments = member.get("lineAssignments") or []
        if line_assignments:
            for assignment in line_assignments:
                member_sheet.append([
                    assignment.get("receiptTitle"),
                    assignment.get("merchantName"),
                    assignment.get("receiptDate"),
                    assignment.get("receiptTime"),
                    assignment.get("paidByMemberName") or "Ikke valgt",
                    assignment.get("itemName"),
                    assignment.get("type"),
                    assignment.get("quantity"),
                    assignment.get("percent"),
                    assignment.get("amount"),
                    assignment.get("unitPrice"),
                    assignment.get("lineTotal"),
                ])
        else:
            member_sheet.append(["Ingen fordelte linjer", "", "", "", "", "", "", "", "", "", "", ""])

    for sheet in workbook.worksheets:
        autosize(sheet)

    workbook.save(output_path)


def table_style(table):
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{ACCENT}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8B9A7")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{LIGHT}")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def build_registered_pdf(payload, output_path):
    result = payload["result"]
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Registrert kostnad", styles["Title"]),
        Spacer(1, 6 * mm),
        Paragraph(text(result.get("merchantName"), "Ukjent sted"), styles["Heading2"]),
        Paragraph(f"Dato: {text(result.get('receiptDate'))}   Tid: {text(result.get('receiptTime'))}", styles["BodyText"]),
        Spacer(1, 4 * mm),
    ]

    summary_rows = [
        ["Kvittering total", money(result.get("grandTotal"))],
        ["Sum registrerte varer", money(result.get("totals", {}).get("itemsTotal"))],
        ["Avvik", money(result.get("totals", {}).get("difference"))],
    ]
    story.append(table_style(Table([["Felt", "Verdi"], *summary_rows], repeatRows=1)))
    story.append(Spacer(1, 6 * mm))

    rows = [["Vare", "Antall", "Pris pr", "Linjesum"]]
    for item in result.get("lineItems") or result.get("items") or []:
        rows.append([
            text(item.get("name")),
            text(item.get("quantity")),
            money(item.get("unitPrice")),
            money(item.get("lineTotal")),
        ])
    story.append(table_style(Table(rows, repeatRows=1)))

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    doc.build(story)


def build_distribution_pdf(payload, output_path):
    participants, remaining_entries, remaining_total = summarize_distribution(payload["distributionState"])
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Fordeling av regning", styles["Title"]),
        Spacer(1, 6 * mm),
        Paragraph(f"Gjenstar a fordele: {money(remaining_total)}", styles["Heading2"]),
        Spacer(1, 4 * mm),
    ]

    participant_rows = [["Person", "Total"]]
    for participant in participants:
        participant_rows.append([participant.get("name"), money(participant.get("total"))])
    story.append(table_style(Table(participant_rows, repeatRows=1)))
    story.append(Spacer(1, 6 * mm))

    remaining_rows = [["Vare", "Gjenstaende antall", "Pris pr", "Gjenstaende sum"]]
    for entry in remaining_entries:
        remaining_rows.append([
            entry.get("name"),
            text(entry.get("remainingQuantity")),
            money(entry.get("unitPrice")),
            money(entry.get("remainingTotal")),
        ])
    story.append(table_style(Table(remaining_rows, repeatRows=1)))

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    doc.build(story)


def build_participant_pdf(payload, output_path):
    participant = resolve_participant(payload)
    participants, _, _ = summarize_distribution(payload["distributionState"])
    participant_summary = next((item for item in participants if item["id"] == participant["id"]), None)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Fordeling for {participant.get('name')}", styles["Title"]),
        Spacer(1, 6 * mm),
        Paragraph(f"Total: {money(participant_summary.get('total') if participant_summary else 0)}", styles["Heading2"]),
        Spacer(1, 4 * mm),
    ]

    rows = [["Vare", "Type", "Antall", "Andel %", "Belop"]]
    for assignment in participant_summary.get("assignments", []) if participant_summary else []:
        rows.append([
            assignment.get("entryName"),
            assignment.get("type"),
            text(assignment.get("quantity")),
            text(assignment.get("percent") or ""),
            money(assignment.get("amount")),
        ])
    story.append(table_style(Table(rows, repeatRows=1)))

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    doc.build(story)


def build_event_pdf(payload, output_path):
    summary_data = payload.get("summary") or {}
    receipts = payload.get("receipts") or []
    members = summary_data.get("members") or []
    styles = getSampleStyleSheet()
    table_cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["BodyText"],
        fontSize=8.5,
        leading=10,
        spaceBefore=0,
        spaceAfter=0,
    )
    story = [
        Paragraph("Arrangementsrapport", styles["Title"]),
        Spacer(1, 6 * mm),
        Paragraph(text(payload.get("eventName"), "Arrangement"), styles["Heading2"]),
        Paragraph("Oppsummering av hele arrangementet med balanse per medlem og kvitteringsdetaljer.", styles["BodyText"]),
        Spacer(1, 4 * mm),
    ]

    summary_rows = [
        ["Kvitteringer", text(summary_data.get("receiptCount") or 0)],
        ["Totalt brukt", money(summary_data.get("totalSpent"))],
        ["Registrert betalt", money(summary_data.get("totalPaid"))],
        ["Fordelt pa personer", money(summary_data.get("totalUsed"))],
        ["Ufordelt rest", money(summary_data.get("unassignedTotal"))],
        ["Mangler betaler", text(summary_data.get("missingPayerCount") or 0)],
    ]
    story.append(table_style(Table([["Felt", "Verdi"], *summary_rows], repeatRows=1, colWidths=[70 * mm, 40 * mm])))
    story.append(Spacer(1, 6 * mm))

    member_rows = [["Person", "Betalt", "Brukt", "Balanse", "Betalte kvitteringer"]]
    if members:
        for member in members:
            member_rows.append([
                text(member.get("name")),
                money(member.get("paidTotal")),
                money(member.get("usedTotal")),
                money(member.get("balance")),
                text(member.get("paidReceiptCount") or 0),
            ])
    else:
        member_rows.append(["Ingen medlemmer ennå", "", "", "", ""])
    story.append(
        table_style(
            Table(
                wrap_table_rows(member_rows, table_cell_style),
                repeatRows=1,
                colWidths=[55 * mm, 28 * mm, 28 * mm, 28 * mm, 35 * mm],
            )
        )
    )
    story.append(Spacer(1, 6 * mm))

    overview_rows = [["Kvittering / sted", "Dato", "Betalt av", "Total", "Fordelt", "Ufordelt"]]
    if receipts:
        for receipt in receipts:
            title = text(receipt.get("title"))
            merchant = text(receipt.get("merchantName"))
            overview_rows.append([
                f"{title}\n{merchant}" if merchant else title,
                " ".join(part for part in [text(receipt.get("receiptDate")), text(receipt.get("receiptTime"))] if part).strip(),
                text(receipt.get("paidByMemberName"), "Ikke valgt"),
                money(receipt.get("grandTotal")),
                money(receipt.get("distributedTotal")),
                money(receipt.get("unassignedTotal")),
            ])
    else:
        overview_rows.append(["Ingen kvitteringer registrert", "", "", "", "", ""])
    story.append(
        table_style(
            Table(
                wrap_table_rows(overview_rows, table_cell_style),
                repeatRows=1,
                colWidths=[58 * mm, 26 * mm, 32 * mm, 20 * mm, 20 * mm, 22 * mm],
            )
        )
    )

    if receipts:
        story.append(Spacer(1, 8 * mm))

    for receipt in receipts:
        story.extend([
            Paragraph(
                text(receipt.get("title"), "Kvittering"),
                styles["Heading3"],
            ),
            Paragraph(
                " | ".join(
                    part
                    for part in [
                        text(receipt.get("merchantName")),
                        text(receipt.get("receiptDate")),
                        text(receipt.get("receiptTime")),
                        text(receipt.get("paidByMemberName"), "Ikke valgt"),
                    ]
                    if part
                ),
                styles["BodyText"],
            ),
            Spacer(1, 3 * mm),
        ])

        receipt_rows = [
            ["Felt", "Verdi"],
            ["Kvittering total", money(receipt.get("grandTotal"))],
            ["Sum registrerte varer", money(receipt.get("itemsTotal"))],
            ["Avvik", money(receipt.get("difference"))],
            ["Fordelt", money(receipt.get("distributedTotal"))],
            ["Ufordelt", money(receipt.get("unassignedTotal"))],
        ]
        story.append(
            table_style(
                Table(
                    wrap_table_rows(receipt_rows, table_cell_style),
                    repeatRows=1,
                    colWidths=[58 * mm, 36 * mm],
                )
            )
        )
        story.append(Spacer(1, 4 * mm))

        item_rows = [["Vare", "Antall", "Pris pr", "Linjesum"]]
        line_items = receipt.get("lineItems") or []
        if line_items:
            for item in line_items:
                item_rows.append([
                    text(item.get("name")),
                    text(item.get("quantity")),
                    money(item.get("unitPrice")),
                    money(item.get("lineTotal")),
                ])
        else:
            item_rows.append(["Ingen linjer registrert", "", "", ""])
        story.append(
            table_style(
                Table(
                    wrap_table_rows(item_rows, table_cell_style),
                    repeatRows=1,
                    colWidths=[92 * mm, 20 * mm, 28 * mm, 28 * mm],
                )
            )
        )
        story.append(Spacer(1, 7 * mm))

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    doc.build(story)


def main():
    _, payload_path, output_path, kind, file_format = sys.argv
    payload = load_payload(payload_path)

    builders = {
        ("registered", "xlsx"): build_registered_xlsx,
        ("registered", "pdf"): build_registered_pdf,
        ("distribution", "xlsx"): build_distribution_xlsx,
        ("distribution", "pdf"): build_distribution_pdf,
        ("participant", "xlsx"): build_participant_xlsx,
        ("participant", "pdf"): build_participant_pdf,
        ("event", "xlsx"): build_event_xlsx,
        ("event", "pdf"): build_event_pdf,
    }

    builder = builders[(kind, file_format)]
    builder(payload, output_path)


if __name__ == "__main__":
    main()
